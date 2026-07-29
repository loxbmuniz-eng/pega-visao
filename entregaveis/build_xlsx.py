# -*- coding: utf-8 -*-
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# ---------- paleta da marca ----------
GRAFITE   = "1A1A1A"
GRAFITE_2 = "2E2E2E"
CREME     = "F5F0E5"
OLIVA     = "A8B848"
OLIVA_D   = "6E7A2A"
MAGENTA   = "D8336B"
MAGENTA_D = "A61F4D"
DOURADO   = "C9A063"
DOURADO_D = "8A6B33"
TEAL_D    = "2C6B7D"
CINZA     = "6B6B6B"
LINHA     = "DCD6C8"
ZEBRA     = "FBF9F4"

FONT = "Arial"
thin = Side(style="thin", color=LINHA)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Pega Visão x UP19"
ws.sheet_view.showGridLines = False

# ---------- dados ----------
TRAZEMOS = [
 ("Obra Pega Visão",
  "Obra tátil original cedida ao festival para exposição na galeria oficial do UP, durante todo o período do evento.",
  None, "Contrapartida"),
 ("Live painting — Badaró", "Pintura ao vivo durante o festival. Cachê solicitado ao UP — ver bloco O QUE PEDIMOS.", None, "Contrapartida"),
 ("Cerâmica Às Cegas — Mickaela",
  "Vivência guiada de olhos vendados, dias 28, 29 e 30/12. Cachê solicitado ao UP — ver bloco O QUE PEDIMOS.", None, "Contrapartida"),
 ("Cobertura audiovisual — Zambonini",
  "Captação e produção de conteúdo documental do projeto no UP19. Cachê solicitado ao UP — ver bloco O QUE PEDIMOS.", None, "Contrapartida"),
 ("Equipamento audiovisual próprio",
  "Kit profissional do diretor: câmera Sony A7IV (33MP), lentes Sigma 24-70 Art e 70-200 GM, lapela Hollyland Lark Max 2, "
  "tripé, iluminação (tubo LED Sokani 25W, LED portátil 40W) e flash Godox V1. Garante padrão profissional de captação das "
  "experiências do Pega Visão e sustenta o intercâmbio de material audiovisual proposto com a equipe oficial do UP.",
  "45,000 em equipamentos", "Qualidade técnica"),
 ("Trust in Dance — Fernanda Pistelli",
  "Experiência de dança guiada de olhos vendados. Participação artística, sem custo ao festival.", None, "Contrapartida"),
 ("Sets de referência",
  "Dois sets autorais de Luis (Ota Blind) enviados para apreciação da curadoria.", None, "Contrapartida"),
 ("Passagens aéreas — equipe",
  "Ida e volta: Luis Otávio, André Zambonini e Luis Badaró. Cotação real em 27/07/2026, ida 26/12/2026 e volta 05/01/2027 — "
  "Cuiabá↔Salvador (Badaró e Zambonini, 2 pax): R$5.317,00. Brasília↔Salvador (Luis, 1 pax): R$2.266,00.",
  7583, "Pega Visão"),
 ("Translado Salvador ↔ Pratigi", "Ida e volta para a equipe (Luis, Badaró, Zambonini).", 930, "Pega Visão"),
 ("Hospedagem da equipe",
  "Casa própria no Jatimane — hospedagem e base de operação durante todo o período.", 7000, "Pega Visão"),
 ("Portfólios e material de referência",
  "Portfólios de Zambonini e Badaró; registros da vivência de cerâmica com pessoas com deficiência visual (edição anterior); "
  "vídeos da experiência Trust in Dance.", None, "Anexo"),
 ("Intercâmbio audiovisual",
  "Material captado pelo Pega Visão disponibilizado ao UP para uso institucional.", None, "Contrapartida"),
]

PEDIMOS = [
 ("Cachê — Live painting (Badaró)", "Cachê de execução do live painting durante o festival.", 300, "Custo a cobrir pelo UP"),
 ("Cachê — Cerâmica Às Cegas (Mickaela)", "Cachê de execução da vivência guiada, dias 28, 29 e 30/12.", 300, "Custo a cobrir pelo UP"),
 ("Cachê — Cobertura audiovisual (Zambonini)", "Cachê de execução da captação documental do projeto no UP19.", 300, "Custo a cobrir pelo UP"),
 ("Materiais — Live painting Badaró",
  "Tinta acrílica, suporte MDF 220×180, embalagem e proteção para transporte.", 980, "Custo a cobrir pelo UP"),
 ("Produção — Nova obra tátil (Cimática)",
  "Materiais para a nova peça do artista Luis Badaró, com tema cimática — formas visuais das frequências sonoras.",
  "A cotar", "Custo a cobrir pelo UP"),
 ("Materiais — Vendas para os olhos",
  "25 metros de tecido para confecção de 300 vendas — higienizadas, personalizadas, utilizadas em todas as experiências do "
  "Pega Visão no UP.", 750, "Custo a cobrir pelo UP"),
 ("Materiais — Cerâmica Às Cegas (Mickaela)", "Argila, lona, bacias.", 500, "Custo a cobrir pelo UP"),
]

APOIO = [
 ("Espaço — Live painting",
  "Área próxima à galeria de arte do UP para execução do live painting ao vivo durante o festival."),
 ("Slot — Galeria de arte",
  "Espaço na galeria oficial do UP para exposição de nova obra tátil, produzida especialmente para o UP19, com tema cimática. "
  "A cimática é o estudo de como o som, em vibração, organiza a matéria em padrões visíveis — a prova física de que toda "
  "frequência sonora carrega uma forma, mesmo antes de qualquer olho a perceber. A obra parte desse princípio: transforma, em "
  "relevo tátil, os padrões que diferentes frequências imprimem na matéria — as figuras que o som desenha quando encontra "
  "areia, água ou placas em vibração. Ela convida quem vê a tocar o que o som desenha, e permite que quem não vê acesse, pela "
  "primeira vez, a forma visual da própria música. É a alma do Pega Visão em sua expressão mais literal: a certeza de que a "
  "percepção do mundo nunca dependeu de um único sentido — o som sempre teve corpo, mesmo para quem nunca pôde vê-lo."),
 ("Cronograma da obra",
  "Definição do dia em que a nova obra precisa estar entregue na galeria e da antecedência necessária para montagem."),
 ("Espaço — Cerâmica Às Cegas",
  "Local para a vivência guiada nos dias 28, 29 e 30/12 — Circulou, Espaço de Cura, ou área próxima ao Chill Out."),
 ("Suporte — Live painting",
  "MDF 220×180 cedido ou produzido localmente pelo UP (a confirmar); definição de fixação — cavalete ou lona/parede."),
 ("Slot — Chill Out", "Set de Ota Blind no Chill Out, em horário próximo à experiência Trust in Dance — horário exato [A CONFIRMAR], conforme grade de shows de Fernanda Pistelli."),
 ("Slot — UP Club", "Set de Ota Blind no UP Club."),
 ("Credenciais — equipe", "3 credenciais: Badaró, Zambonini e 1 vaga de apoio operacional."),
 ("Credencial — artista", "Luis (Ota Blind): credencial de artista + 1 acompanhante."),
 ("Acesso a mídias — UP",
  "Acesso ao material audiovisual captado pela produção oficial do UP, como parte do intercâmbio proposto."),
]

APORTE_COR = {
    "Pega Visão": OLIVA_D, "Contrapartida": TEAL_D, "Qualidade técnica": DOURADO_D,
    "Anexo": CINZA, "Custo a cobrir pelo UP": MAGENTA_D,
}

# ---------- larguras ----------
for col, w in (("A", 31), ("B", 68), ("C", 20), ("D", 23)):
    ws.column_dimensions[col].width = w

MOEDA = '[$-416]R$ #,##0.00'          # LCID pt-BR: força R$ 16.113,00 em qualquer locale

def alt(row, cols=4):
    fill = PatternFill("solid", fgColor=ZEBRA if row % 2 == 0 else "FFFFFF")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill

def altura(textos, larguras, minimo=26):
    linhas = 1
    for t, w in zip(textos, larguras):
        if t:
            linhas = max(linhas, math.ceil(len(str(t)) / (w * 1.05)))
    return max(minimo, linhas * 12.2 + 7)

def faixa(row, titulo, cor_fundo, cor_texto, span="A{r}:D{r}"):
    ws.merge_cells(span.format(r=row))
    c = ws.cell(row=row, column=1, value=titulo)
    c.font = Font(name=FONT, size=11.5, bold=True, color=cor_texto)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i in range(1, 5):
        ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=cor_fundo)
    ws.row_dimensions[row].height = 27

def cabecalho(row, labels, cols=(1, 2, 3, 4), merges=None):
    for label, col in zip(labels, cols):
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color=CREME)
        c.alignment = Alignment(horizontal="right" if label.startswith("Valor") else "left",
                                vertical="center", indent=1)
    for i in range(1, 5):
        ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=GRAFITE_2)
    for m in (merges or []):
        ws.merge_cells(m)
    ws.row_dimensions[row].height = 23

r = 1
# ===== TÍTULO =====
ws.merge_cells("A1:D1")
t = ws.cell(row=1, column=1, value="PEGA VISÃO × UNIVERSO PARALELLO 19")
t.font = Font(name=FONT, size=19, bold=True, color=CREME)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
for i in range(1, 5):
    ws.cell(row=1, column=i).fill = PatternFill("solid", fgColor=GRAFITE)
ws.row_dimensions[1].height = 50

ws.merge_cells("A2:D2")
s = ws.cell(row=2, column=1,
            value="Provisionamento  ·  Equipe: Luis, Badaró, Zambonini  ·  Micka e Fernanda com participação e logística próprias")
s.font = Font(name=FONT, size=9.5, color=DOURADO)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
for i in range(1, 5):
    ws.cell(row=2, column=i).fill = PatternFill("solid", fgColor=GRAFITE)
ws.row_dimensions[2].height = 24
ws.row_dimensions[3].height = 8

# ===== BLOCO 1 =====
faixa(4, "O QUE TRAZEMOS", OLIVA, GRAFITE)
cabecalho(5, ["Item", "Descrição", "Valor (R$)", "Aporte"])
r = 6
primeira_trazemos = r
for item, desc, valor, aporte in TRAZEMOS:
    alt(r)
    a = ws.cell(row=r, column=1, value=item)
    a.font = Font(name=FONT, size=10, bold=True, color=GRAFITE)
    a.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    b = ws.cell(row=r, column=2, value=desc)
    b.font = Font(name=FONT, size=9, color="333333")
    b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    c = ws.cell(row=r, column=3, value=valor if valor is not None else "—")
    if isinstance(valor, (int, float)):
        c.number_format = MOEDA
        c.font = Font(name=FONT, size=10.5, bold=True, color=GRAFITE)
    else:
        c.font = Font(name=FONT, size=9.5, italic=True, color=CINZA)
    c.alignment = Alignment(horizontal="right", vertical="top", indent=1, wrap_text=True)
    d = ws.cell(row=r, column=4, value=aporte)
    d.font = Font(name=FONT, size=9, bold=True, color=APORTE_COR[aporte])
    d.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = BORDER
    ws.row_dimensions[r].height = altura([item, desc], [31, 68])
    r += 1
ultima_trazemos = r - 1

ws.row_dimensions[r].height = 8
r += 1
linha_total_1 = r
ws.merge_cells(f"A{r}:B{r}")
c = ws.cell(row=r, column=1, value="TOTAL — APORTE PEGA VISÃO")
c.font = Font(name=FONT, size=12, bold=True, color=CREME)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
v = ws.cell(row=r, column=3, value=f"=SUM(C{primeira_trazemos}:C{ultima_trazemos})")
v.number_format = MOEDA
v.font = Font(name=FONT, size=13.5, bold=True, color=OLIVA)
v.alignment = Alignment(horizontal="right", vertical="center", indent=1)
for i in range(1, 5):
    ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GRAFITE)
ws.row_dimensions[r].height = 34
r += 1
ws.row_dimensions[r].height = 16
r += 1

# ===== BLOCO 2 =====
faixa(r, "O QUE PEDIMOS  ·  CUSTO A COBRIR PELO UP", MAGENTA, "FFFFFF"); r += 1
cabecalho(r, ["Item", "Descrição", "Valor (R$)", "Natureza"]); r += 1
primeira_pedimos = r
for item, desc, valor, nat in PEDIMOS:
    alt(r)
    a = ws.cell(row=r, column=1, value=item)
    a.font = Font(name=FONT, size=10, bold=True, color=GRAFITE)
    a.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    b = ws.cell(row=r, column=2, value=desc)
    b.font = Font(name=FONT, size=9, color="333333")
    b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    c = ws.cell(row=r, column=3, value=valor)
    if isinstance(valor, (int, float)):
        c.number_format = MOEDA
        c.font = Font(name=FONT, size=10.5, bold=True, color=GRAFITE)
    else:
        c.font = Font(name=FONT, size=9.5, italic=True, bold=True, color=MAGENTA_D)
    c.alignment = Alignment(horizontal="right", vertical="top", indent=1, wrap_text=True)
    d = ws.cell(row=r, column=4, value=nat)
    d.font = Font(name=FONT, size=9, bold=True, color=MAGENTA_D)
    d.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = BORDER
    ws.row_dimensions[r].height = altura([item, desc], [31, 68])
    r += 1
ultima_pedimos = r - 1

ws.row_dimensions[r].height = 8
r += 1
ws.merge_cells(f"A{r}:B{r}")
c = ws.cell(row=r, column=1, value="TOTAL — SOLICITADO AO UP")
c.font = Font(name=FONT, size=12, bold=True, color=CREME)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
v = ws.cell(row=r, column=3, value=f"=SUM(C{primeira_pedimos}:C{ultima_pedimos})")
v.number_format = MOEDA
v.font = Font(name=FONT, size=13.5, bold=True, color=MAGENTA)
v.alignment = Alignment(horizontal="right", vertical="center", indent=1)
for i in range(1, 5):
    ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GRAFITE)
ws.row_dimensions[r].height = 34
r += 1
ws.merge_cells(f"A{r}:D{r}")
n = ws.cell(row=r, column=1, value="+ valor da produção da obra tátil cimática, a definir")
n.font = Font(name=FONT, size=9.5, italic=True, color=MAGENTA_D)
n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
for i in range(1, 5):
    ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor="FBEEF3")
ws.row_dimensions[r].height = 22
linha_nota_2 = r
r += 1
ws.row_dimensions[r].height = 16
r += 1

# ===== BLOCO 3 =====
faixa(r, "O QUE PEDIMOS  ·  APOIO INSTITUCIONAL — SEM VALOR MONETÁRIO", DOURADO, GRAFITE); r += 1
cabecalho(r, ["Item", "Descrição — apoio institucional"], cols=(1, 2), merges=[f"B{r}:D{r}"]); r += 1
for item, desc in APOIO:
    alt(r)
    ws.merge_cells(f"B{r}:D{r}")
    a = ws.cell(row=r, column=1, value=item)
    a.font = Font(name=FONT, size=10, bold=True, color=GRAFITE)
    a.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    b = ws.cell(row=r, column=2, value=desc)
    b.font = Font(name=FONT, size=9, color="333333")
    b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = BORDER
    ws.row_dimensions[r].height = altura([item, desc], [31, 112])
    r += 1

# quebras de página nos limites de bloco (cada bloco inteiro na sua página)
for br in (linha_total_1, linha_nota_2):
    ws.row_breaks.append(Break(id=br))

ws.freeze_panes = "A6"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.print_title_rows = "1:2"

out = "/home/user/pega-visao/entregaveis/PEGA_VISAO_x_UP19_PROVISIONAMENTO.xlsx"
wb.save(out)
print("salvo:", out, "| ultima linha:", r - 1)
